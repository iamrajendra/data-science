#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Function to take multiple arguments
import  openpyxl as xl
# upload the spreedsheet
wb  = xl.load_workbook('transections.xlsx')
sheet = wb['Sheet1']
cell  = sheet['a1']
cell  = sheet.cell(1,1)

for row  in range(1,sheet.max_row+1):
    cell  = sheet.cell(row,3)
    corrected_price  = cell.value*0.9
    corrected_price_cell   = sheet.cell(cell,4)
    corrected_price_cell.value  = corrected_price_cell
wb.save("transection2.xlsx")
